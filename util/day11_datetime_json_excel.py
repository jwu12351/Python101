import time
from datetime import datetime, timedelta
import openpyxl
import json


if __name__ == '__main__':
    # 1970.01.01 以来的秒数
    t = time.time()
    print(type(t))
    print(t)
    print()

    # time.struct_time(tm_year=2023, tm_mon=8, tm_mday=18, tm_hour=0, tm_min=7, tm_sec=12,
    # tm_wday=4, tm_yday=230, tm_isdst=0)
    t = time.localtime()
    print(type(t))
    print(t)

    for x in iter(t):
        print(x, end=' ')
    print('\n')


    # %y 两位数的年份表示（00-99）
    # %Y 四位数的年份表示（000-9999）
    # %m 月份（01-12）
    # %d 月内中的一天（1-31）

    # %H 24小时制小时数（0-23）
    # %I 12小时制小时数（01-12）
    # %M 分钟数（00-59）
    # %S 秒（00-59）

    t = time.asctime()
    print(type(t))
    print(t)
    print()

    # today is 2023.08.18 00:14:15 !
    t = time.strftime('today is %Y.%m.%d %H:%M:%S !', time.localtime())
    print(type(t))
    print(t)
    print()

    now = datetime.now()
    print(type(now))
    print(now)

    now = datetime.now()
    print(now)
    print(now.strftime('%Y'), now.strftime('%m'), now.strftime('%d'))
    print(now.strftime('%H'), now.strftime('%M'), now.strftime('%S'))
    print()

    date_time = now.strftime("%Y-%m-%d, %H:%M:%S")
    print(date_time)

    tomorrow = now + timedelta(days=1)
    yesterday = now + timedelta(days=-1)
    print(tomorrow)
    print(yesterday)
    print()

    date = datetime.strptime('2023-03-04 10:45:00', "%Y-%m-%d %H:%M:%S")
    print(type(date))
    print(date)
    print()

    # 计算天数
    start_date = "2023-01-01"
    end_date = "2023-06-14"
    start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
    end_datetime = datetime.strptime(end_date, "%Y-%m-%d")
    delta = end_datetime - start_datetime
    print(type(delta))
    days = delta.days
    print(days)
    print()


    # JSON
    # JavaScript Object Notation
    # 作用：一种数据交换格式
    #   数字、字符串、布尔值
    #   列表、词典
    # json.dumps(): 对 json对象 编码 成 字符串
    # json.loads(): 对 字符串 解码成 json对象
    # json.dump(): 对 json文件数据 编码成字符串
    # json.load(): 对 文件数据 解码成json对象

    people_info = {'name': 'xxx', 'info': [1, 2, 3, {'a': 1, 'b': 2}]}
    value = json.dumps(people_info)
    print(type(value))
    print(value)

    json_object = json.loads(value)
    print(type(json_object))
    print(json_object)

    with open('people.json', 'w+') as f:
        json.dump(people_info, f)
    print()

    with open('people.json', 'r+') as f:
        value = json.load(f)
        print(type(value))
        print(value)


# Excel
# openpyxl
# pip install openpyxl
def read_xlsx_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.worksheets[0]
    # sheet = workbook.active

    rows = sheet.max_row
    cols = sheet.max_column

    for i in range(1, rows + 1):
        for j in range(1, cols + 1):
            sheet.cell(i, j, '11')
            print(sheet.cell(i, j).value, end=' ')
        print()

    workbook.save('result.xlsx')


read_xlsx_file('test.xlsx')
